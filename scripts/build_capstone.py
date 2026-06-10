from __future__ import annotations

import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW_PATH = PROJECT_ROOT / "data" / "raw" / "Delhi_NCR_Delivery_Operations_Raw.csv"
OUTPUT_DIR = PROJECT_ROOT / "outputs"
PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"
DOCS_DIR = PROJECT_ROOT / "docs"
POWERBI_DIR = PROJECT_ROOT / "powerbi"


def parse_rating(value: object) -> float:
    if pd.isna(value):
        return np.nan
    text = str(value).strip().lower()
    rating_map = {
        "terrible": 1.0,
        "poor": 2.0,
        "average": 3.0,
        "good": 4.0,
        "excellent": 5.0,
    }
    if text in rating_map:
        return rating_map[text]
    match = re.search(r"\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else np.nan


def money(value: float) -> str:
    return f"Rs {value:,.0f}"


def pct(value: float) -> str:
    return f"{value:.1%}"


def clean_orders() -> pd.DataFrame:
    df = pd.read_csv(RAW_PATH)
    df = df.drop_duplicates(subset=["Order_ID"], keep="first").copy()
    df["Order_Timestamp"] = pd.to_datetime(df["Order_Timestamp"], errors="coerce")
    df["Order_Date"] = df["Order_Timestamp"].dt.date
    df["Order_Hour"] = df["Order_Timestamp"].dt.hour
    df["Day_Name"] = df["Order_Timestamp"].dt.day_name()
    df["Week"] = df["Order_Timestamp"].dt.isocalendar().week.astype("Int64")
    df["Rider_Rating"] = df["Rider_Rating_Raw"].apply(parse_rating)
    df["Delivery_Speed_KMPH"] = (
        df["Delivery_Distance_KM"] / (df["Transit_Time_Mins"] / 60)
    ).replace([np.inf, -np.inf], np.nan)
    df["Revenue_Per_KM"] = (df["Order_Value_INR"] / df["Delivery_Distance_KM"]).replace(
        [np.inf, -np.inf], np.nan
    )
    df["Delivery_Bucket"] = pd.cut(
        df["Total_Delivery_Time_Mins"],
        bins=[0, 10, 15, 20, 30, np.inf],
        labels=["0-10", "10-15", "15-20", "20-30", "30+"],
        right=False,
    )
    df["Distance_Bucket"] = pd.cut(
        df["Delivery_Distance_KM"],
        bins=[0, 1, 2, 3, 5, np.inf],
        labels=["0-1 km", "1-2 km", "2-3 km", "3-5 km", "5+ km"],
        right=False,
    )
    df = df.drop(columns=["Legacy_Marketing_ID", "Temp_System_Log", "Customer_Phone_Raw"])
    df = df.dropna(subset=["Order_ID", "Order_Timestamp", "Dark_Store_ID", "SLA_Breached"])
    df["Total_Items"] = df["Total_Items"].astype("int64")
    df["Order_Value_INR"] = df["Order_Value_INR"].astype("float64")
    df["SLA_Breached"] = df["SLA_Breached"].astype("int64")
    df["Order_Hour"] = df["Order_Hour"].astype("int64")
    df["Week"] = df["Week"].astype("int64")
    return df


def aggregate_outputs(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    store = (
        df.groupby("Dark_Store_ID")
        .agg(
            Orders=("Order_ID", "count"),
            Revenue=("Order_Value_INR", "sum"),
            Avg_Order_Value=("Order_Value_INR", "mean"),
            Avg_Delivery_Time=("Total_Delivery_Time_Mins", "mean"),
            SLA_Breach_Rate=("SLA_Breached", "mean"),
            Avg_Distance=("Delivery_Distance_KM", "mean"),
            Avg_Rider_Rating=("Rider_Rating", "mean"),
        )
        .reset_index()
        .sort_values("SLA_Breach_Rate", ascending=False)
    )

    hourly = (
        df.groupby("Order_Hour")
        .agg(
            Orders=("Order_ID", "count"),
            Revenue=("Order_Value_INR", "sum"),
            Avg_Delivery_Time=("Total_Delivery_Time_Mins", "mean"),
            SLA_Breach_Rate=("SLA_Breached", "mean"),
        )
        .reset_index()
    )

    driver = (
        df.groupby(["Traffic_Density", "Weather", "Vehicle_Type"])
        .agg(
            Orders=("Order_ID", "count"),
            Avg_Delivery_Time=("Total_Delivery_Time_Mins", "mean"),
            SLA_Breach_Rate=("SLA_Breached", "mean"),
            Avg_Rider_Rating=("Rider_Rating", "mean"),
        )
        .reset_index()
        .sort_values(["SLA_Breach_Rate", "Orders"], ascending=[False, False])
    )

    category = (
        df.groupby("Primary_Category")
        .agg(
            Orders=("Order_ID", "count"),
            Revenue=("Order_Value_INR", "sum"),
            Avg_Order_Value=("Order_Value_INR", "mean"),
            SLA_Breach_Rate=("SLA_Breached", "mean"),
            Avg_Delivery_Time=("Total_Delivery_Time_Mins", "mean"),
        )
        .reset_index()
        .sort_values("Revenue", ascending=False)
    )

    local_rider = (
        df.groupby("Is_Local_Rider")
        .agg(
            Orders=("Order_ID", "count"),
            Revenue=("Order_Value_INR", "sum"),
            Avg_Order_Value=("Order_Value_INR", "mean"),
            Avg_Delivery_Time=("Total_Delivery_Time_Mins", "mean"),
            SLA_Breach_Rate=("SLA_Breached", "mean"),
            Avg_Distance=("Delivery_Distance_KM", "mean"),
            Avg_Rider_Rating=("Rider_Rating", "mean"),
        )
        .reset_index()
        .sort_values("Is_Local_Rider", ascending=False)
    )

    sla_matrix = pd.pivot_table(
        df,
        values="SLA_Breached",
        index="Traffic_Density",
        columns="Distance_Bucket",
        aggfunc="mean",
    ).reset_index()

    return {
        "store_performance": store,
        "hourly_trend": hourly,
        "delivery_drivers": driver,
        "category_performance": category,
        "local_rider_impact": local_rider,
        "sla_matrix": sla_matrix,
    }


def write_csv_outputs(df: pd.DataFrame, tables: dict[str, pd.DataFrame]) -> None:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(PROCESSED_DIR / "orders_clean.csv", index=False)
    for name, table in tables.items():
        table.to_csv(PROCESSED_DIR / f"{name}.csv", index=False)


def add_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start_col):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, start_col):
            if pd.isna(value):
                value = None
            ws.cell(row_idx, col_idx, value)
    ws.freeze_panes = ws.cell(start_row + 1, start_col).coordinate
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, start_col):
        max_len = max(
            len(str(header)),
            *(len(str(ws.cell(row, col_idx).value or "")) for row in range(start_row + 1, min(ws.max_row, 120) + 1)),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 28)


def style_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_excel(df: pd.DataFrame, tables: dict[str, pd.DataFrame], metrics: dict[str, object]) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    ws["A1"] = "Delhi-NCR Quick Commerce Optimization Capstone"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    ws["A3"] = "Business Objective"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = "Reduce SLA breaches, improve delivery speed, and identify store/rider/operational levers."
    kpis = [
        ("Total Orders", metrics["orders"]),
        ("Total Revenue", money(metrics["revenue"])),
        ("Average Delivery Time", f"{metrics['avg_delivery_time']:.2f} mins"),
        ("SLA Breach Rate", pct(metrics["sla_rate"])),
        ("Average Order Value", money(metrics["aov"])),
        ("Average Rider Rating", f"{metrics['avg_rating']:.2f} / 5"),
    ]
    for idx, (label, value) in enumerate(kpis, 5):
        ws.cell(idx, 1, label).font = Font(bold=True)
        ws.cell(idx, 2, value)
    insights = [
        f"Highest SLA risk store: {metrics['worst_store']}",
        f"Peak order hour: {metrics['peak_hour']}:00",
        f"Highest revenue category: {metrics['top_category']}",
        f"Local riders reduce SLA breach rate by {metrics['local_rider_sla_improvement_pp']:.2f} percentage points.",
        f"Recommended focus: dispatch capacity during high-traffic, long-distance windows.",
    ]
    ws["A13"] = "Key Insights"
    ws["A13"].font = Font(bold=True)
    for row_idx, insight in enumerate(insights, 14):
        ws.cell(row_idx, 1, insight)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 42
    style_sheet(ws)

    for sheet_name, table in [
        ("Store Performance", tables["store_performance"]),
        ("Hourly Trend", tables["hourly_trend"]),
        ("Delivery Drivers", tables["delivery_drivers"].head(100)),
        ("Category Performance", tables["category_performance"]),
        ("Local Rider Impact", tables["local_rider_impact"]),
        ("SLA Matrix", tables["sla_matrix"]),
        ("Clean Sample", df.head(1000)),
    ]:
        sheet = wb.create_sheet(sheet_name)
        add_dataframe(sheet, table)
        style_sheet(sheet)

    store_ws = wb["Store Performance"]
    store_ws.conditional_formatting.add(
        f"E2:E{store_ws.max_row}",
        ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
    )

    hour_ws = wb["Hourly Trend"]
    chart = LineChart()
    chart.title = "Orders by Hour"
    chart.y_axis.title = "Orders"
    chart.x_axis.title = "Hour"
    data = Reference(hour_ws, min_col=2, min_row=1, max_row=hour_ws.max_row)
    cats = Reference(hour_ws, min_col=1, min_row=2, max_row=hour_ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    hour_ws.add_chart(chart, "G2")

    cat_ws = wb["Category Performance"]
    bar = BarChart()
    bar.title = "Revenue by Category"
    bar.y_axis.title = "Revenue"
    bar.x_axis.title = "Category"
    data = Reference(cat_ws, min_col=3, min_row=1, max_row=cat_ws.max_row)
    cats = Reference(cat_ws, min_col=1, min_row=2, max_row=cat_ws.max_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 8
    bar.width = 16
    cat_ws.add_chart(bar, "H2")

    dd = wb.create_sheet("Power BI Blueprint")
    rows = [
        ["Page", "Visual", "Purpose"],
        ["Executive Overview", "KPI cards", "Orders, revenue, AOV, SLA breach rate, average delivery time"],
        ["Operations Control", "Store bar chart", "Rank stores by SLA breach rate and delivery time"],
        ["Demand Planning", "Hourly line chart", "Identify peak slots for staffing and dispatch planning"],
        ["Root Cause", "Matrix heatmap", "Compare SLA risk by traffic, weather, vehicle, and distance"],
        ["Recommendations", "Text cards", "Translate findings into actions for city operations"],
    ]
    for row in rows:
        dd.append(row)
    add_dataframe(dd, pd.DataFrame(rows[1:], columns=rows[0]))
    style_sheet(dd)

    path = OUTPUT_DIR / "Delhi_NCR_Quick_Commerce_Capstone_Cleaned.xlsx"
    wb.save(path)
    return path


def write_docs(metrics: dict[str, object]) -> None:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    POWERBI_DIR.mkdir(parents=True, exist_ok=True)

    readme = f"""# Delhi-NCR Quick Commerce Optimization Capstone

## Objective
Build an end-to-end analytics project to reduce delivery SLA breaches and improve operational efficiency for quick commerce deliveries across Delhi-NCR.

## Tools Used
- Python: data cleaning, feature engineering, KPI aggregation, repeatable pipeline
- Excel: executive summary workbook, KPI tables, charts, and operational analysis tabs
- Power BI: dashboard model design, DAX measures, and report-building blueprint

## Dataset
- Raw file: `Delhi_NCR_Delivery_Operations_Raw.csv`
- Rows analyzed: {metrics["orders"]:,}
- Duplicate orders removed during cleaning: {metrics["duplicate_orders_removed"]:,}
- Revenue analyzed: {money(metrics["revenue"])}
- Overall SLA breach rate: {pct(metrics["sla_rate"])}
- Local rider SLA breach rate: {pct(metrics["local_rider_sla_rate"])}
- Non-local rider SLA breach rate: {pct(metrics["non_local_rider_sla_rate"])}
- Local rider SLA improvement: {metrics["local_rider_sla_improvement_pp"]:.2f} percentage points

## Business Questions
1. Which stores contribute most to SLA breaches?
2. Which delivery conditions create the highest operational risk?
3. When should staffing and rider allocation be increased?
4. Which categories and order profiles drive revenue?
5. What operational actions can reduce late deliveries?
6. Do local riders improve delivery speed and SLA performance?

## Recommended Power BI Pages
1. Executive Overview
2. Store Performance
3. Delivery SLA Root Cause
4. Demand and Staffing Plan
5. Recommendations

## Final Recommendation
Prioritize local riders for high-risk delivery zones, increase dispatch capacity for high-traffic and long-distance orders, monitor underperforming dark stores daily, and create rider allocation rules around peak order hours.
"""
    (DOCS_DIR / "README.md").write_text(readme, encoding="utf-8")

    dax = """Orders = COUNTROWS(orders_clean)
Revenue = SUM(orders_clean[Order_Value_INR])
Average Order Value = DIVIDE([Revenue], [Orders])
SLA Breaches = SUM(orders_clean[SLA_Breached])
SLA Breach Rate = DIVIDE([SLA Breaches], [Orders])
Average Delivery Time = AVERAGE(orders_clean[Total_Delivery_Time_Mins])
Average Pack Time = AVERAGE(orders_clean[Pack_Time_Mins])
Average Transit Time = AVERAGE(orders_clean[Transit_Time_Mins])
Average Distance = AVERAGE(orders_clean[Delivery_Distance_KM])
Average Rider Rating = AVERAGE(orders_clean[Rider_Rating])
Revenue Per KM = DIVIDE([Revenue], SUM(orders_clean[Delivery_Distance_KM]))
"""
    (POWERBI_DIR / "dax_measures.txt").write_text(dax, encoding="utf-8")

    guide = """# Power BI Build Guide

## Load Data
Import `data/processed/orders_clean.csv` into Power BI Desktop.

## Model
Use `Order_Date`, `Order_Hour`, `Dark_Store_ID`, `Weather`, `Traffic_Density`, `Vehicle_Type`, `Primary_Category`, and `Distance_Bucket` as slicer fields.

## Measures
Create the measures listed in `powerbi/dax_measures.txt`.

## Dashboard Pages
- Executive Overview: KPI cards, revenue by category, hourly order trend
- Store Performance: SLA breach rate by dark store, average delivery time, revenue
- Root Cause: SLA rate by traffic, weather, vehicle type, and distance bucket
- Demand Planning: orders by hour and day, peak period staffing recommendation
- Recommendations: operational actions and expected business impact
"""
    (POWERBI_DIR / "powerbi_build_guide.md").write_text(guide, encoding="utf-8")


def main() -> None:
    df = clean_orders()
    tables = aggregate_outputs(df)
    raw_df = pd.read_csv(RAW_PATH)
    metrics = {
        "orders": int(len(df)),
        "raw_orders": int(len(raw_df)),
        "duplicate_orders_removed": int(raw_df.duplicated("Order_ID").sum()),
        "revenue": float(df["Order_Value_INR"].sum()),
        "avg_delivery_time": float(df["Total_Delivery_Time_Mins"].mean()),
        "sla_rate": float(df["SLA_Breached"].mean()),
        "aov": float(df["Order_Value_INR"].mean()),
        "avg_rating": float(df["Rider_Rating"].mean()),
        "worst_store": str(tables["store_performance"].iloc[0]["Dark_Store_ID"]),
        "peak_hour": int(tables["hourly_trend"].sort_values("Orders", ascending=False).iloc[0]["Order_Hour"]),
        "top_category": str(tables["category_performance"].iloc[0]["Primary_Category"]),
    }
    local_rider = tables["local_rider_impact"].set_index("Is_Local_Rider")
    metrics["local_rider_avg_delivery_time"] = float(local_rider.loc["Yes", "Avg_Delivery_Time"])
    metrics["non_local_rider_avg_delivery_time"] = float(local_rider.loc["No", "Avg_Delivery_Time"])
    metrics["local_rider_sla_rate"] = float(local_rider.loc["Yes", "SLA_Breach_Rate"])
    metrics["non_local_rider_sla_rate"] = float(local_rider.loc["No", "SLA_Breach_Rate"])
    metrics["local_rider_sla_improvement_pp"] = (
        metrics["non_local_rider_sla_rate"] - metrics["local_rider_sla_rate"]
    ) * 100
    metrics["local_rider_time_saving_mins"] = (
        metrics["non_local_rider_avg_delivery_time"] - metrics["local_rider_avg_delivery_time"]
    )
    write_csv_outputs(df, tables)
    workbook_path = build_excel(df, tables, metrics)
    write_docs(metrics)
    summary_path = OUTPUT_DIR / "project_summary.json"
    summary_path.write_text(json.dumps(metrics, indent=2), encoding="utf-8")
    print(f"Workbook: {workbook_path}")
    print(f"Clean data: {PROCESSED_DIR / 'orders_clean.csv'}")
    print(json.dumps(metrics, indent=2))


if __name__ == "__main__":
    main()
